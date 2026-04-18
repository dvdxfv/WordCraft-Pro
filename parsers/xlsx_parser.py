"""
XLSX 解析器

解析 .xlsx 文件，将每个 sheet 转换为表格数据。
"""

from __future__ import annotations

from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from parsers.base import BaseParser
from core.document_model import (
    DocumentModel, DocElement, ElementType, ParagraphStyle, Alignment,
    TableData, TableCell,
)


class XlsxParser(BaseParser):
    """Excel (.xlsx) 文件解析器"""

    supported_extensions = [".xlsx"]

    def parse(self, file_path: str, **kwargs) -> DocumentModel:
        """解析 .xlsx 文件"""
        abs_path = self._validate_file(file_path)
        wb = load_workbook(abs_path, read_only=True, data_only=True)

        model = DocumentModel(
            source_file=abs_path,
            source_format="xlsx",
            title=kwargs.get("title", ""),
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 添加 sheet 名称作为标题
            model.elements.append(DocElement(
                element_type=ElementType.HEADING,
                content=f"工作表: {sheet_name}",
                level=2,
                font_style=__import__("core.document_model", fromlist=["FontStyle"]).FontStyle(bold=True),
            ))

            # 解析表格数据
            table = self._parse_sheet(ws, sheet_name)
            if table and table.num_rows > 0:
                model.tables.append(table)

                # 添加表格标题元素
                model.elements.append(DocElement(
                    element_type=ElementType.CAPTION,
                    content=table.caption,
                    paragraph_style=ParagraphStyle(alignment=Alignment.CENTER),
                ))

        wb.close()
        return model

    def _parse_sheet(self, ws, sheet_name: str) -> Optional[TableData]:
        """解析单个工作表为 TableData"""
        rows = []
        max_col = 0

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            # 跳过完全空的行
            if all(cell is None for cell in row):
                continue

            cells = []
            for col_idx, value in enumerate(row, start=1):
                cell_text = str(value) if value is not None else ""
                is_header = (row_idx == 1)

                cells.append(TableCell(
                    content=cell_text.strip(),
                    is_header=is_header,
                ))

            if cells:
                rows.append(cells)
                max_col = max(max_col, len(cells))

        if not rows:
            return None

        return TableData(
            caption=f"表: {sheet_name}",
            numbering="",
            rows=rows,
        )
